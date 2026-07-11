"""
Money-Pro Excel 模板生成器
从 ledger.csv 读取数据，生成带图表的 Excel 展示文件
"""

import csv
import os
from datetime import datetime
from pathlib import Path
from collections import defaultdict

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


# ============================================================
# 配置
# ============================================================

CATEGORIES = {
    "车": ["过路费", "能源费用", "停车费", "车贷月供", "保险", "车载用品购物", "洗车美容", "保养维修"],
    "宠物": ["主粮零食", "玩具", "医疗与洗护"],
    "日常生活": ["餐饮", "购物", "医疗", "月供", "其他"],
    "收入": ["工资", "奖金", "兼职", "理财收益", "红包", "其他"],
}

MONTHS = ["1月", "2月", "3月", "4月", "5月", "6月",
          "7月", "8月", "9月", "10月", "11月", "12月"]

CURRENT_YEAR = datetime.now().year

# 样式定义
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(name="微软雅黑", size=10, bold=True)
ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
SUMMARY_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
CATEGORY_FONT = Font(name="微软雅黑", size=10, bold=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 类别颜色
CATEGORY_COLORS = {
    "车": "3498DB",
    "宠物": "E74C3C",
    "日常生活": "2ECC71",
    "收入": "27AE60",
}

# 记录人颜色
PERSON_COLORS = ["4472C4", "ED7D31"]  # 蓝色、橙色


# ============================================================
# 数据加载
# ============================================================

def load_ledger(csv_path):
    """从 ledger.csv 加载数据"""
    records = []
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                # 解析日期
                date_str = row.get('日期', '')
                if date_str:
                    date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                else:
                    continue
                
                # 解析金额
                amount = float(row.get('金额', 0))
                
                records.append({
                    'date': date_obj,
                    'date_str': date_str,
                    'month': date_obj.month,
                    'category': row.get('类别', '其他'),
                    'subcategory': row.get('子类别', '未分类'),
                    'amount': amount,
                    'direction': row.get('收支类型', '支出'),
                    'payment': row.get('支付方式', '其他'),
                    'person': row.get('记录人', '未知'),
                    'note': row.get('备注', ''),
                    'source': row.get('来源', '手动')
                })
            except Exception as e:
                continue
    
    return records


# ============================================================
# Excel 创建
# ============================================================

def create_workbook():
    """创建工作簿"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb


def create_instruction_sheet(wb, persons):
    """创建使用说明Sheet"""
    ws = wb.create_sheet("使用说明")
    ws.sheet_properties.tabColor = "1F4E79"
    
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 50
    
    # 标题
    ws.merge_cells('B2:C2')
    ws['B2'] = "Money-Pro 日常生活记账表格"
    ws['B2'].font = Font(name="微软雅黑", size=18, bold=True, color="1F4E79")
    
    ws.merge_cells('B3:C3')
    ws['B3'] = f"版本 2.0 | {CURRENT_YEAR}年 | 双人记账版"
    ws['B3'].font = Font(name="微软雅黑", size=10, color="808080")
    
    # 功能介绍
    row = 5
    ws.merge_cells(f'B{row}:C{row}')
    ws[f'B{row}'] = "一、系统架构"
    ws[f'B{row}'].font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")
    
    row = 6
    features = [
        "1. 数据源: ledger.csv (唯一真相，支持 Git 版本控制)",
        "2. 展示层: 本 Excel 文件 (从 ledger.csv 读取数据)",
        f"3. 记录人: {', '.join(persons)}",
        "4. 导入工具: import.py (支持微信/支付宝账单导入)",
    ]
    for i, feature in enumerate(features):
        ws[f'B{row+i}'] = feature
        ws[f'B{row+i}'].font = Font(name="微软雅黑", size=10)
    
    # 使用流程
    row = 11
    ws.merge_cells(f'B{row}:C{row}')
    ws[f'B{row}'] = "二、日常记账流程"
    ws[f'B{row}'].font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")
    
    row = 12
    steps = [
        "方式一: 手动记账",
        "  1. 打开 ledger.csv 文件",
        "  2. 在末尾添加新行，填写日期、类别、金额等字段",
        "  3. 保存 CSV 文件",
        "  4. 运行 python refresh_excel.py 刷新 Excel 数据",
        "",
        "方式二: 导入账单",
        "  1. 从微信/支付宝导出账单文件",
        "  2. 运行导入命令:",
        f"     python import.py wechat '微信账单.xlsx' --person {persons[0]}",
        f"     python import.py alipay '支付宝账单.csv' --person {persons[1]}",
        "  3. 运行 python refresh_excel.py 刷新 Excel 数据",
    ]
    for step in steps:
        ws[f'B{row}'] = step
        ws[f'B{row}'].font = Font(name="微软雅黑", size=10)
        row += 1
    
    # 类别说明
    row += 1
    ws.merge_cells(f'B{row}:C{row}')
    ws[f'B{row}'] = "三、收支类别说明"
    ws[f'B{row}'].font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")
    
    row += 1
    ws[f'B{row}'] = "类别"
    ws[f'C{row}'] = "子类别"
    ws[f'B{row}'].font = SUBHEADER_FONT
    ws[f'C{row}'].font = SUBHEADER_FONT
    ws[f'B{row}'].fill = SUBHEADER_FILL
    ws[f'C{row}'].fill = SUBHEADER_FILL
    
    row += 1
    for cat, subs in CATEGORIES.items():
        ws[f'B{row}'] = cat
        ws[f'C{row}'] = "、".join(subs)
        ws[f'B{row}'].font = CATEGORY_FONT
        color = CATEGORY_COLORS.get(cat, "808080")
        ws[f'B{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws[f'B{row}'].font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        row += 1
    
    # 快捷键
    row += 1
    ws.merge_cells(f'B{row}:C{row}')
    ws[f'B{row}'] = "四、常用快捷键"
    ws[f'B{row}'].font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")
    
    row += 1
    shortcuts = [
        ("Ctrl + ;", "插入当前日期"),
        ("Ctrl + Shift + ;", "插入当前时间"),
        ("Ctrl + D", "向下填充"),
        ("Ctrl + Enter", "在选中区域输入相同内容"),
        ("Alt + Enter", "单元格内换行"),
    ]
    for key, desc in shortcuts:
        ws[f'B{row}'] = key
        ws[f'C{row}'] = desc
        ws[f'B{row}'].font = Font(name="Consolas", size=10)
        row += 1
    
    ws.freeze_panes = 'A1'
    return ws


def create_data_sheet(wb, records, persons):
    """创建数据透视Sheet"""
    ws = wb.create_sheet("数据透视")
    ws.sheet_properties.tabColor = "1F4E79"
    
    # 列宽
    col_widths = {'A': 12, 'B': 10, 'C': 12, 'D': 12, 'E': 10,
                  'F': 10, 'G': 10, 'H': 25, 'I': 10}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    # 标题
    ws.merge_cells('A1:I1')
    ws['A1'] = f"Money-Pro - {CURRENT_YEAR}年数据总览"
    ws['A1'].font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # 筛选器说明
    ws.merge_cells('A3:I3')
    ws['A3'] = "提示: 可使用 Excel 自带的筛选功能按记录人、月份、类别等筛选数据"
    ws['A3'].font = Font(name="微软雅黑", size=10, color="808080")
    
    # 表头
    headers = ["日期", "类别", "子类别", "金额", "收支类型", "支付方式", "记录人", "备注", "来源"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
    
    ws.row_dimensions[4].height = 25
    
    # 数据行
    # 按日期排序
    sorted_records = sorted(records, key=lambda x: x['date'], reverse=True)
    
    for i, record in enumerate(sorted_records):
        row = 5 + i
        
        # 日期
        ws.cell(row=row, column=1, value=record['date_str']).border = THIN_BORDER
        ws.cell(row=row, column=1).number_format = 'yyyy-mm-dd'
        
        # 类别
        ws.cell(row=row, column=2, value=record['category']).border = THIN_BORDER
        
        # 子类别
        ws.cell(row=row, column=3, value=record['subcategory']).border = THIN_BORDER
        
        # 金额
        cell_amount = ws.cell(row=row, column=4, value=record['amount'])
        cell_amount.number_format = '#,##0.00'
        cell_amount.border = THIN_BORDER
        
        # 收支类型
        ws.cell(row=row, column=5, value=record['direction']).border = THIN_BORDER
        
        # 支付方式
        ws.cell(row=row, column=6, value=record['payment']).border = THIN_BORDER
        
        # 记录人
        ws.cell(row=row, column=7, value=record['person']).border = THIN_BORDER
        
        # 备注
        ws.cell(row=row, column=8, value=record['note']).border = THIN_BORDER
        
        # 来源
        ws.cell(row=row, column=9, value=record['source']).border = THIN_BORDER
        
        # 交替行颜色
        if i % 2 == 1:
            for col in range(1, 10):
                ws.cell(row=row, column=col).fill = ALT_ROW_FILL
    
    # 添加自动筛选
    ws.auto_filter.ref = f"A4:I{4 + len(sorted_records)}"
    
    ws.freeze_panes = 'A5'
    return ws


def create_monthly_sheet(wb, month_num, month_name, records, persons):
    """创建月度明细表"""
    ws = wb.create_sheet(month_name)
    ws.sheet_properties.tabColor = CATEGORY_COLORS.get("餐饮", "FF8C00")
    
    # 筛选当月数据
    month_records = [r for r in records if r['month'] == month_num]
    
    # 列宽
    col_widths = {'A': 12, 'B': 10, 'C': 12, 'D': 20, 'E': 12,
                  'F': 10, 'G': 10, 'H': 8, 'I': 3,
                  'J': 10, 'K': 12, 'L': 12, 'M': 12, 'N': 8}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    # 标题
    ws.merge_cells('A1:H1')
    ws['A1'] = f"Money-Pro - {CURRENT_YEAR}年{month_name}支出记录"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # 表头
    headers = ["日期", "类别", "子类别", "备注", "金额(元)", "支付方式", "收支类型", "记录人"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
    
    ws.row_dimensions[2].height = 25
    
    # 数据行
    sorted_records = sorted(month_records, key=lambda x: x['date'], reverse=True)
    
    for i, record in enumerate(sorted_records):
        row = 3 + i
        
        ws.cell(row=row, column=1, value=record['date_str']).border = THIN_BORDER
        ws.cell(row=row, column=1).number_format = 'yyyy-mm-dd'
        
        ws.cell(row=row, column=2, value=record['category']).border = THIN_BORDER
        ws.cell(row=row, column=3, value=record['subcategory']).border = THIN_BORDER
        ws.cell(row=row, column=4, value=record['note']).border = THIN_BORDER
        
        cell_amount = ws.cell(row=row, column=5, value=record['amount'])
        cell_amount.number_format = '#,##0.00'
        cell_amount.border = THIN_BORDER
        
        ws.cell(row=row, column=6, value=record['payment']).border = THIN_BORDER
        ws.cell(row=row, column=7, value=record['direction']).border = THIN_BORDER
        ws.cell(row=row, column=8, value=record['person']).border = THIN_BORDER
        
        if i % 2 == 1:
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = ALT_ROW_FILL

    data_rows = max(len(sorted_records), 1)

    # 添加自动筛选
    if len(sorted_records) > 0:
        ws.auto_filter.ref = f"A2:H{2 + len(sorted_records)}"
    
    # ========================================
    # 月度汇总区域（J-N列）
    # ========================================
    summary_col_start = 10  # J列
    
    # 汇总标题
    ws.merge_cells(f'J1:N1')
    ws['J1'] = "月度汇总"
    ws['J1'].font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")
    ws['J1'].fill = SUMMARY_FILL
    ws['J1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 汇总表头
    summary_headers = ["类别", persons[0], persons[1], "合计", "占比"]
    for col_idx, header in enumerate(summary_headers):
        cell = ws.cell(row=2, column=summary_col_start + col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
    
    # 汇总数据
    category_list = list(CATEGORIES.keys())
    
    # 按类别和记录人统计（包含支出和收入）
    cat_person_amount = defaultdict(lambda: defaultdict(float))
    for record in month_records:
        cat_person_amount[record['category']][record['person']] += record['amount']
    
    expense_cats = [c for c in category_list if c != '收入']
    income_cat = '收入'
    
    # 各记录人支出/收入汇总
    expense_by_person = defaultdict(float)
    income_by_person = defaultdict(float)
    
    for i, cat in enumerate(category_list):
        row = 3 + i
        
        # 类别名称
        cell_j = ws.cell(row=row, column=summary_col_start, value=cat)
        color = CATEGORY_COLORS.get(cat, "808080")
        cell_j.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell_j.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        cell_j.border = THIN_BORDER
        cell_j.alignment = Alignment(horizontal='center')
        
        # 各记录人金额
        for p_idx, person in enumerate(persons):
            amount = cat_person_amount[cat].get(person, 0)
            if cat != income_cat:
                expense_by_person[person] += amount
            else:
                income_by_person[person] += amount
            cell = ws.cell(row=row, column=summary_col_start + 1 + p_idx, value=amount)
            cell.number_format = '#,##0.00'
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
        
        # 合计
        cat_total = sum(cat_person_amount[cat].values())
        cell = ws.cell(row=row, column=summary_col_start + 3, value=cat_total)
        cell.number_format = '#,##0.00'
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        
        # 占比（稍后计算）
        ws.cell(row=row, column=summary_col_start + 4).border = THIN_BORDER
    
    # 支出总计行
    expense_total = sum(expense_by_person.values())
    total_row = 3 + len(category_list)
    ws.cell(row=total_row, column=summary_col_start, value="支出总计").font = Font(
        name="微软雅黑", size=10, bold=True)
    ws.cell(row=total_row, column=summary_col_start).fill = SUMMARY_FILL
    ws.cell(row=total_row, column=summary_col_start).border = THIN_BORDER
    
    for p_idx, person in enumerate(persons):
        cell = ws.cell(row=total_row, column=summary_col_start + 1 + p_idx, 
                       value=expense_by_person[person])
        cell.number_format = '#,##0.00'
        cell.font = Font(name="微软雅黑", size=10, bold=True)
        cell.fill = SUMMARY_FILL
        cell.border = THIN_BORDER
    
    ws.cell(row=total_row, column=summary_col_start + 3, value=expense_total)
    ws.cell(row=total_row, column=summary_col_start + 3).number_format = '#,##0.00'
    ws.cell(row=total_row, column=summary_col_start + 3).font = Font(
        name="微软雅黑", size=10, bold=True)
    ws.cell(row=total_row, column=summary_col_start + 3).fill = SUMMARY_FILL
    ws.cell(row=total_row, column=summary_col_start + 3).border = THIN_BORDER
    
    ws.cell(row=total_row, column=summary_col_start + 4, value="100%")
    ws.cell(row=total_row, column=summary_col_start + 4).font = Font(
        name="微软雅黑", size=10, bold=True)
    ws.cell(row=total_row, column=summary_col_start + 4).fill = SUMMARY_FILL
    ws.cell(row=total_row, column=summary_col_start + 4).border = THIN_BORDER
    
    # 收入总计行
    income_total = sum(income_by_person.values())
    income_row = total_row + 1
    ws.cell(row=income_row, column=summary_col_start, value="收入总计").font = Font(
        name="微软雅黑", size=10, bold=True)
    ws.cell(row=income_row, column=summary_col_start).fill = PatternFill(
        start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    ws.cell(row=income_row, column=summary_col_start).border = THIN_BORDER
    
    for p_idx, person in enumerate(persons):
        cell = ws.cell(row=income_row, column=summary_col_start + 1 + p_idx,
                       value=income_by_person[person])
        cell.number_format = '#,##0.00'
        cell.font = Font(name="微软雅黑", size=10, bold=True, color="27AE60")
        cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        cell.border = THIN_BORDER
    
    ws.cell(row=income_row, column=summary_col_start + 3, value=income_total)
    ws.cell(row=income_row, column=summary_col_start + 3).number_format = '#,##0.00'
    ws.cell(row=income_row, column=summary_col_start + 3).font = Font(
        name="微软雅黑", size=10, bold=True, color="27AE60")
    ws.cell(row=income_row, column=summary_col_start + 3).fill = PatternFill(
        start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    ws.cell(row=income_row, column=summary_col_start + 3).border = THIN_BORDER
    
    # 结余行
    net_row = income_row + 1
    net_amount = income_total - expense_total
    net_color = "27AE60" if net_amount >= 0 else "E74C3C"
    ws.cell(row=net_row, column=summary_col_start, value="本月结余").font = Font(
        name="微软雅黑", size=10, bold=True)
    ws.cell(row=net_row, column=summary_col_start).fill = PatternFill(
        start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
    ws.cell(row=net_row, column=summary_col_start).border = THIN_BORDER
    
    for p_idx, person in enumerate(persons):
        p_net = income_by_person[person] - expense_by_person[person]
        p_color = "27AE60" if p_net >= 0 else "E74C3C"
        cell = ws.cell(row=net_row, column=summary_col_start + 1 + p_idx, value=p_net)
        cell.number_format = '#,##0.00'
        cell.font = Font(name="微软雅黑", size=10, bold=True, color=p_color)
        cell.fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
        cell.border = THIN_BORDER
    
    cell = ws.cell(row=net_row, column=summary_col_start + 3, value=net_amount)
    cell.number_format = '#,##0.00'
    cell.font = Font(name="微软雅黑", size=11, bold=True, color=net_color)
    cell.fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
    cell.border = THIN_BORDER
    
    # 回填占比（仅支出类别）
    for i, cat in enumerate(expense_cats):
        row = 3 + category_list.index(cat)
        cat_total = ws.cell(row=row, column=summary_col_start + 3).value or 0
        if expense_total > 0:
            pct = cat_total / expense_total
        else:
            pct = 0
        cell = ws.cell(row=row, column=summary_col_start + 4, value=pct)
        cell.number_format = '0.0%'
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
    
    # ========================================
    # 图表
    # ========================================

    # 饼图 - 支出占比（仅支出类别）
    cat_end_row = 3 + len(expense_cats) - 1
    pie = PieChart()
    pie.title = f"{month_name}支出占比"
    pie.style = 10
    pie.width = 18
    pie.height = 14

    labels = Reference(ws, min_col=summary_col_start, min_row=3,
                       max_row=cat_end_row)
    data = Reference(ws, min_col=summary_col_start + 3, min_row=2,
                     max_row=cat_end_row)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = True
    pie.dataLabels.showCatName = True

    for i, cat in enumerate(expense_cats):
        pt = DataPoint(idx=i)
        color = CATEGORY_COLORS.get(cat, "808080")
        pt.graphicalProperties.solidFill = color
        pie.series[0].data_points.append(pt)

    chart_start_row = total_row + 3
    ws.add_chart(pie, f'J{chart_start_row}')

    # 水平柱状图 - 各类别支出金额（仅支出类别）
    bar = BarChart()
    bar.type = "bar"
    bar.title = f"{month_name} 各类别支出金额"
    bar.style = 10
    bar.width = 20
    bar.height = 14
    bar.x_axis.title = "金额(元)"

    # 数据来源：合计列
    data = Reference(ws, min_col=summary_col_start + 3, min_row=2,
                     max_row=cat_end_row)
    bar.add_data(data, titles_from_data=True)

    # 类别作为Y轴标签
    cats = Reference(ws, min_col=summary_col_start, min_row=3,
                     max_row=cat_end_row)
    bar.set_categories(cats)

    # 添加数据标签
    bar.dataLabels = DataLabelList()
    bar.dataLabels.showVal = True
    bar.dataLabels.numFmt = '#,##0.00'

    # 设置每根柱子颜色
    for i, cat in enumerate(expense_cats):
        pt = DataPoint(idx=i)
        color = CATEGORY_COLORS.get(cat, "808080")
        pt.graphicalProperties.solidFill = color
        bar.series[0].data_points.append(pt)

    ws.add_chart(bar, f'J{chart_start_row + 16}')

    ws.freeze_panes = 'A3'
    return ws


def create_person_summary_sheet(wb, records, persons):
    """创建按人汇总Sheet"""
    ws = wb.create_sheet("按人汇总")
    ws.sheet_properties.tabColor = "4472C4"
    
    # 列宽
    for col in range(1, 16):
        ws.column_dimensions[get_column_letter(col)].width = 12
    
    # 标题
    ws.merge_cells('A1:H1')
    ws['A1'] = f"Money-Pro - {CURRENT_YEAR}年按人汇总"
    ws['A1'].font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
    ws.row_dimensions[1].height = 30
    
    # ========================================
    # 支出表
    # ========================================
    expense_headers = ["月份"] + persons + ["支出合计"]
    for col_idx, h in enumerate(expense_headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
    
    # 按月统计支出
    expense_data = defaultdict(lambda: defaultdict(float))
    income_data = defaultdict(lambda: defaultdict(float))
    for record in records:
        if record['direction'] == '支出':
            expense_data[record['month']][record['person']] += record['amount']
        elif record['direction'] == '收入':
            income_data[record['month']][record['person']] += record['amount']
    
    for month_idx in range(12):
        row = 4 + month_idx
        month_num = month_idx + 1
        
        ws.cell(row=row, column=1, value=MONTHS[month_idx]).border = THIN_BORDER
        ws.cell(row=row, column=1).font = Font(name="微软雅黑", size=10, bold=True)
        
        month_total = 0
        for p_idx, person in enumerate(persons):
            amount = expense_data[month_num].get(person, 0)
            month_total += amount
            cell = ws.cell(row=row, column=p_idx + 2, value=amount)
            cell.number_format = '#,##0.00'
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
            # 红色标出大额支出
            if amount > 500:
                cell.font = Font(name="微软雅黑", size=10, color="E74C3C")
        
        cell = ws.cell(row=row, column=len(persons) + 2, value=month_total)
        cell.number_format = '#,##0.00'
        cell.font = Font(name="微软雅黑", size=10, bold=True)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        
        if month_idx % 2 == 1:
            for col in range(1, len(persons) + 3):
                ws.cell(row=row, column=col).fill = ALT_ROW_FILL
    
    # 支出年度总计
    exp_total_row = 16
    ws.cell(row=exp_total_row, column=1, value="支出总计").font = Font(
        name="微软雅黑", size=10, bold=True)
    ws.cell(row=exp_total_row, column=1).fill = SUMMARY_FILL
    ws.cell(row=exp_total_row, column=1).border = THIN_BORDER
    
    exp_grand_by_person = defaultdict(float)
    for p_idx, person in enumerate(persons):
        col = p_idx + 2
        total = sum(expense_data[m+1].get(person, 0) for m in range(12))
        exp_grand_by_person[person] = total
        cell = ws.cell(row=exp_total_row, column=col, value=total)
        cell.number_format = '#,##0.00'
        cell.font = Font(name="微软雅黑", size=10, bold=True)
        cell.fill = SUMMARY_FILL
        cell.border = THIN_BORDER
    
    exp_grand_total = sum(exp_grand_by_person.values())
    cell = ws.cell(row=exp_total_row, column=len(persons) + 2, value=exp_grand_total)
    cell.number_format = '#,##0.00'
    cell.font = Font(name="微软雅黑", size=11, bold=True, color="FF0000")
    cell.fill = SUMMARY_FILL
    cell.border = THIN_BORDER
    
    # ========================================
    # 收入表（追加在支出表下方）
    # ========================================
    inc_header_row = exp_total_row + 2
    inc_headers = ["月份"] + persons + ["收入合计"]
    for col_idx, h in enumerate(inc_headers, 1):
        cell = ws.cell(row=inc_header_row, column=col_idx, value=h)
        cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
    
    for month_idx in range(12):
        row = inc_header_row + 1 + month_idx
        month_num = month_idx + 1
        
        ws.cell(row=row, column=1, value=MONTHS[month_idx]).border = THIN_BORDER
        ws.cell(row=row, column=1).font = Font(name="微软雅黑", size=10, bold=True)
        
        month_total = 0
        for p_idx, person in enumerate(persons):
            amount = income_data[month_num].get(person, 0)
            month_total += amount
            cell = ws.cell(row=row, column=p_idx + 2, value=amount)
            cell.number_format = '#,##0.00'
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
        
        cell = ws.cell(row=row, column=len(persons) + 2, value=month_total)
        cell.number_format = '#,##0.00'
        cell.font = Font(name="微软雅黑", size=10, bold=True)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        
        if month_idx % 2 == 1:
            for col in range(1, len(persons) + 3):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    
    # 收入年度总计
    inc_footer_row = inc_header_row + 13
    ws.cell(row=inc_footer_row, column=1, value="收入总计").font = Font(
        name="微软雅黑", size=10, bold=True)
    ws.cell(row=inc_footer_row, column=1).fill = PatternFill(
        start_color="A5D6A7", end_color="A5D6A7", fill_type="solid")
    ws.cell(row=inc_footer_row, column=1).border = THIN_BORDER
    
    inc_grand_by_person = defaultdict(float)
    for p_idx, person in enumerate(persons):
        col = p_idx + 2
        total = sum(income_data[m+1].get(person, 0) for m in range(12))
        inc_grand_by_person[person] = total
        cell = ws.cell(row=inc_footer_row, column=col, value=total)
        cell.number_format = '#,##0.00'
        cell.font = Font(name="微软雅黑", size=10, bold=True, color="27AE60")
        cell.fill = PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid")
        cell.border = THIN_BORDER
    
    inc_grand_total = sum(inc_grand_by_person.values())
    cell = ws.cell(row=inc_footer_row, column=len(persons) + 2, value=inc_grand_total)
    cell.number_format = '#,##0.00'
    cell.font = Font(name="微软雅黑", size=11, bold=True, color="27AE60")
    cell.fill = PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid")
    cell.border = THIN_BORDER
    
    # 净收支行
    net_row = inc_footer_row + 1
    ws.cell(row=net_row, column=1, value="年度结余").font = Font(
        name="微软雅黑", size=10, bold=True)
    ws.cell(row=net_row, column=1).fill = PatternFill(
        start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
    ws.cell(row=net_row, column=1).border = THIN_BORDER
    
    for p_idx, person in enumerate(persons):
        col = p_idx + 2
        net_amt = inc_grand_by_person[person] - exp_grand_by_person[person]
        net_color = "27AE60" if net_amt >= 0 else "E74C3C"
        cell = ws.cell(row=net_row, column=col, value=net_amt)
        cell.number_format = '#,##0.00'
        cell.font = Font(name="微软雅黑", size=10, bold=True, color=net_color)
        cell.fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
        cell.border = THIN_BORDER
    
    net_total = inc_grand_total - exp_grand_total
    net_color = "27AE60" if net_total >= 0 else "E74C3C"
    cell = ws.cell(row=net_row, column=len(persons) + 2, value=net_total)
    cell.number_format = '#,##0.00'
    cell.font = Font(name="微软雅黑", size=11, bold=True, color=net_color)
    cell.fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
    cell.border = THIN_BORDER
    
    # ========================================
    # 图表
    # ========================================
    chart_start_row = net_row + 2
    
    # 双折线图 - 两人月度支出趋势
    line = LineChart()
    line.title = f"{CURRENT_YEAR}年月度支出趋势 ({', '.join(persons)})"
    line.style = 10
    line.width = 25
    line.height = 15
    line.y_axis.title = "金额(元)"
    line.x_axis.title = "月份"
    
    for p_idx, person in enumerate(persons):
        data = Reference(ws, min_col=p_idx + 2, min_row=3, max_row=15)
        line.add_data(data, titles_from_data=True)
        line.series[p_idx].graphicalProperties.solidFill = PERSON_COLORS[p_idx]
        line.series[p_idx].graphicalProperties.line.solidFill = PERSON_COLORS[p_idx]
    
    cats = Reference(ws, min_col=1, min_row=4, max_row=15)
    line.set_categories(cats)
    
    ws.add_chart(line, f'A{chart_start_row}')
    
    # 饼图 - 年度各类别总支出占比（仅支出类别）
    cat_totals = defaultdict(float)
    for record in records:
        if record['direction'] == '支出':
            cat_totals[record['category']] += record['amount']
    
    temp_row = chart_start_row + 18
    ws.cell(row=temp_row, column=1, value="类别").font = SUBHEADER_FONT
    ws.cell(row=temp_row, column=2, value="金额").font = SUBHEADER_FONT
    
    category_list = list(CATEGORIES.keys())
    expense_cats = [c for c in category_list if c != '收入']
    for i, cat in enumerate(expense_cats):
        ws.cell(row=temp_row + 1 + i, column=1, value=cat)
        ws.cell(row=temp_row + 1 + i, column=2, value=cat_totals.get(cat, 0))
    
    pie = PieChart()
    pie.title = f"{CURRENT_YEAR}年各类别总支出占比"
    pie.style = 10
    pie.width = 18
    pie.height = 14
    
    labels = Reference(ws, min_col=1, min_row=temp_row + 1, 
                       max_row=temp_row + len(expense_cats))
    data = Reference(ws, min_col=2, min_row=temp_row,
                     max_row=temp_row + len(expense_cats))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    
    for i, cat in enumerate(expense_cats):
        pt = DataPoint(idx=i)
        color = CATEGORY_COLORS.get(cat, "808080")
        pt.graphicalProperties.solidFill = color
        pie.series[0].data_points.append(pt)
    
    ws.add_chart(pie, f'A{chart_start_row + 17}')
    
    ws.freeze_panes = 'B4'
    return ws


def create_yearly_summary_sheet(wb, records, persons):
    """创建年度汇总Sheet"""
    ws = wb.create_sheet("年度汇总")
    ws.sheet_properties.tabColor = "1F4E79"
    
    # 列宽
    col_widths = {'A': 10, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 12,
                  'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 12, 'L': 12,
                  'M': 12, 'N': 12, 'O': 12}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    # 标题
    ws.merge_cells('A1:N1')
    ws['A1'] = f"Money-Pro - {CURRENT_YEAR}年度支出汇总"
    ws['A1'].font = Font(name="微软雅黑", size=16, bold=True, color="1F4E79")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    
    # 月度汇总表
    ws['A3'] = "类别"
    ws['A3'].font = HEADER_FONT
    ws['A3'].fill = HEADER_FILL
    ws['A3'].border = THIN_BORDER
    
    for i, month in enumerate(MONTHS):
        cell = ws.cell(row=3, column=i + 2, value=month)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
    
    cell = ws.cell(row=3, column=14, value="年度合计")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center')
    cell.border = THIN_BORDER
    
    # 按类别和月份统计（包含支出和收入）
    cat_month_amount = defaultdict(lambda: defaultdict(float))
    for record in records:
        cat_month_amount[record['category']][record['month']] += record['amount']
    
    category_list = list(CATEGORIES.keys())
    expense_cats = [c for c in category_list if c != '收入']
    
    for i, cat in enumerate(category_list):
        row = 4 + i
        cell_a = ws.cell(row=row, column=1, value=cat)
        color = CATEGORY_COLORS.get(cat, "808080")
        cell_a.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell_a.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        cell_a.border = THIN_BORDER
        cell_a.alignment = Alignment(horizontal='center')
        
        cat_total = 0
        for j in range(12):
            month_num = j + 1
            amount = cat_month_amount[cat].get(month_num, 0)
            cat_total += amount
            cell = ws.cell(row=row, column=j + 2, value=amount)
            cell.number_format = '#,##0.00'
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
            if i % 2 == 1:
                cell.fill = ALT_ROW_FILL
        
        cell = ws.cell(row=row, column=14, value=cat_total)
        cell.number_format = '#,##0.00'
        cell.font = Font(name="微软雅黑", size=10, bold=True)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
    
    # 月度支出总计行
    exp_total_row = 4 + len(category_list)
    ws.cell(row=exp_total_row, column=1, value="月度支出总计").font = Font(
        name="微软雅黑", size=10, bold=True)
    ws.cell(row=exp_total_row, column=1).fill = SUMMARY_FILL
    ws.cell(row=exp_total_row, column=1).border = THIN_BORDER
    
    for j in range(12):
        month_num = j + 1
        month_total = sum(cat_month_amount[cat].get(month_num, 0) for cat in expense_cats)
        cell = ws.cell(row=exp_total_row, column=j + 2, value=month_total)
        cell.number_format = '#,##0.00'
        cell.font = Font(name="微软雅黑", size=10, bold=True)
        cell.fill = SUMMARY_FILL
        cell.border = THIN_BORDER
    
    exp_grand_total = sum(sum(cat_month_amount[cat].values()) for cat in expense_cats)
    cell = ws.cell(row=exp_total_row, column=14, value=exp_grand_total)
    cell.number_format = '#,##0.00'
    cell.font = Font(name="微软雅黑", size=11, bold=True, color="FF0000")
    cell.fill = SUMMARY_FILL
    cell.border = THIN_BORDER
    
    # 月度收入总计行
    inc_total_row = exp_total_row + 1
    ws.cell(row=inc_total_row, column=1, value="月度收入总计").font = Font(
        name="微软雅黑", size=10, bold=True)
    ws.cell(row=inc_total_row, column=1).fill = PatternFill(
        start_color="A5D6A7", end_color="A5D6A7", fill_type="solid")
    ws.cell(row=inc_total_row, column=1).border = THIN_BORDER
    
    for j in range(12):
        month_num = j + 1
        income_total = cat_month_amount.get('收入', {}).get(month_num, 0)
        cell = ws.cell(row=inc_total_row, column=j + 2, value=income_total)
        cell.number_format = '#,##0.00'
        cell.font = Font(name="微软雅黑", size=10, bold=True, color="27AE60")
        cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        cell.border = THIN_BORDER
    
    inc_grand_total = sum(cat_month_amount.get('收入', {}).values())
    cell = ws.cell(row=inc_total_row, column=14, value=inc_grand_total)
    cell.number_format = '#,##0.00'
    cell.font = Font(name="微软雅黑", size=11, bold=True, color="27AE60")
    cell.fill = PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid")
    cell.border = THIN_BORDER
    
    # 月度结余行
    net_row = inc_total_row + 1
    ws.cell(row=net_row, column=1, value="月度结余").font = Font(
        name="微软雅黑", size=10, bold=True)
    ws.cell(row=net_row, column=1).fill = PatternFill(
        start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
    ws.cell(row=net_row, column=1).border = THIN_BORDER
    
    for j in range(12):
        month_num = j + 1
        exp_amt = sum(cat_month_amount[cat].get(month_num, 0) for cat in expense_cats)
        inc_amt = cat_month_amount.get('收入', {}).get(month_num, 0)
        net_amt = inc_amt - exp_amt
        color = "27AE60" if net_amt >= 0 else "E74C3C"
        cell = ws.cell(row=net_row, column=j + 2, value=net_amt)
        cell.number_format = '#,##0.00'
        cell.font = Font(name="微软雅黑", size=10, bold=True, color=color)
        cell.fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
        cell.border = THIN_BORDER
    
    net_grand_total = inc_grand_total - exp_grand_total
    net_color = "27AE60" if net_grand_total >= 0 else "E74C3C"
    cell = ws.cell(row=net_row, column=14, value=net_grand_total)
    cell.number_format = '#,##0.00'
    cell.font = Font(name="微软雅黑", size=11, bold=True, color=net_color)
    cell.fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
    cell.border = THIN_BORDER
    
    # ========================================
    # 图表
    # ========================================
    chart_start_row = net_row + 2
    
    # 折线图 - 月度支出趋势
    line = LineChart()
    line.title = f"{CURRENT_YEAR}年月度支出趋势"
    line.style = 10
    line.width = 25
    line.height = 15
    line.y_axis.title = "金额(元)"
    line.x_axis.title = "月份"
    
    data = Reference(ws, min_col=2, min_row=exp_total_row,
                     max_col=13, max_row=exp_total_row)
    cats = Reference(ws, min_col=2, min_row=3, max_col=13, max_row=3)
    line.add_data(data, from_rows=True, titles_from_data=False)
    line.set_categories(cats)
    line.series[0].name = "月度总支出"
    
    ws.add_chart(line, f'A{chart_start_row}')
    
    # 堆叠柱状图 - 各类别月度对比（仅支出类别）
    bar_stacked = BarChart()
    bar_stacked.title = f"{CURRENT_YEAR}年各类别支出对比"
    bar_stacked.style = 10
    bar_stacked.width = 25
    bar_stacked.height = 15
    bar_stacked.y_axis.title = "金额(元)"
    bar_stacked.type = "col"
    bar_stacked.grouping = "stacked"
    
    for i, cat in enumerate(expense_cats):
        cat_idx = category_list.index(cat)
        data = Reference(ws, min_col=2, min_row=4 + cat_idx,
                         max_col=13, max_row=4 + cat_idx)
        bar_stacked.add_data(data, titles_from_data=False)
        bar_stacked.series[i].title = openpyxl.chart.series.SeriesLabel(v=cat)
        color = CATEGORY_COLORS.get(cat, "808080")
        bar_stacked.series[i].graphicalProperties.solidFill = color
    
    cats = Reference(ws, min_col=2, min_row=3, max_col=13, max_row=3)
    bar_stacked.set_categories(cats)
    
    ws.add_chart(bar_stacked, f'A{chart_start_row + 17}')
    
    # 年度饼图 - 各类别总支出占比（仅支出类别）
    pie = PieChart()
    pie.title = f"{CURRENT_YEAR}年各类别总支出占比"
    pie.style = 10
    pie.width = 18
    pie.height = 14
    
    labels = Reference(ws, min_col=1, min_row=4, max_row=4 + len(expense_cats) - 1)
    data = Reference(ws, min_col=14, min_row=3, max_row=4 + len(expense_cats) - 1)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    
    for i, cat in enumerate(expense_cats):
        pt = DataPoint(idx=i)
        color = CATEGORY_COLORS.get(cat, "808080")
        pt.graphicalProperties.solidFill = color
        pie.series[0].data_points.append(pt)
    
    ws.add_chart(pie, f'A{chart_start_row + 34}')
    
    ws.freeze_panes = 'B4'
    return ws


# ============================================================
# 主函数
# ============================================================

def main():
    print("开始生成 Money-Pro Excel 模板...")
    
    # 加载配置
    import yaml
    config_path = Path(__file__).parent / "config.yaml"
    with open(config_path, 'r', encoding='utf-8') as f:
        config = yaml.safe_load(f)
    
    persons = config.get('persons', ['李湘云', '粟嘉明'])
    
    # 加载数据
    csv_path = Path(__file__).parent / "ledger.csv"
    if not csv_path.exists():
        print(f"错误: 找不到 ledger.csv")
        return
    
    records = load_ledger(csv_path)
    print(f"从 ledger.csv 加载了 {len(records)} 条记录")
    
    # 创建工作簿
    wb = create_workbook()
    
    # 创建各个Sheet
    print("创建使用说明...")
    create_instruction_sheet(wb, persons)
    
    print("创建数据透视...")
    create_data_sheet(wb, records, persons)
    
    print("创建月度明细表...")
    for i, month in enumerate(MONTHS, 1):
        create_monthly_sheet(wb, i, month, records, persons)
    
    print("创建按人汇总...")
    create_person_summary_sheet(wb, records, persons)
    
    print("创建年度汇总...")
    create_yearly_summary_sheet(wb, records, persons)
    
    # 保存文件
    output_path = Path(__file__).parent / "日常记账表格.xlsx"
    wb.save(output_path)
    print(f"\n文件已保存到: {output_path}")
    print("生成完成！")


if __name__ == '__main__':
    main()
